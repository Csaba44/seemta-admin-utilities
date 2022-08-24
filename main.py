import time
import openpyxl
from datetime import datetime
import time, os


print("Version 4.0")
print()

dutyperc = 0
asegit = 0
setarmor = 0
fix = 0
oilchange = 0
startTime = 0
endTime = 0

adminnick = False
logsfolder = False

f_mtalogs = open("mtalogs.txt", "r")
logsfolder = f_mtalogs.read() + "\console.log"
f_adminnick = open("adminnick.txt", "r")
adminnick = f_adminnick.read()
print("                  oo .d888b. ")
print("                     Y8' `8P ")
print(".d8888b. dP   .dP dP d8bad8b ")
print("88'  `88 88   d8' 88 88` `88 ")
print("88.  .88 88 .88'  88 8b. .88 ")
print("`88888P8 8888P'   dP Y88888P")







try:
    workbook = openpyxl.load_workbook("./admin.xlsx")
    sheet = workbook.active
    sheet["A1"] = adminnick + " ADMIN SHEET"
    sheet["A4"] = "Dutypercek:"
    sheet["A6"] = "Asegit használat:"
    sheet["A8"] = "Setarmor használat:"
    sheet["D4"] = "Fixveh használat:"
    sheet["D6"] = "Oilchange használat:"
    
    if sheet.cell(row = 4, column = 2).value == None:
        dutyperc = 0
        sheet["B4"] = 0
    else:
        dutyperc = sheet.cell(row = 4, column = 2).value
    if sheet.cell(row = 6, column = 2).value == None:
        asegit = 0
        sheet["B6"] = 0
    else:
        asegit = sheet.cell(row = 6, column = 2).value
    if sheet.cell(row = 8, column = 2).value == None:
        setarmor = 0
        sheet["B8"] = 0
    else:
        setarmor = sheet.cell(row = 8, column = 2).value
    if sheet.cell(row = 4, column = 5).value == None:
        fix = 0
        sheet["E4"] = 0
    else:
        fix = sheet.cell(row = 4, column = 5).value
    if sheet.cell(row = 6, column = 5).value == None:
        oilchange = 0
        sheet["E6"] = 0
    else:
        oilchange = sheet.cell(row = 6, column = 5).value

    print("Percek: " + str(dutyperc))
    print("Asegit: " + str(asegit))
    print("Setarmor: " + str(setarmor))
    print("fix: " + str(fix))
    print("oilchange: " + str(oilchange))
    workbook.save("./admin.xlsx")
except:
    print("FAULT: admin.xlsx not found. Create admin.xlsx at the base folder of this script!")

file = open(logsfolder,'r', encoding='utf-8')

st_results = os.stat(logsfolder)
st_size = st_results[6]
file.seek(st_size)

while 1:
    where = file.tell()
    line = file.readline()
    if not line:
        time.sleep(1)
        file.seek(where)
    else:
        print(line) 
        
        
        if "Sikeresen megjavítottad" in line:
            fix += 1
            print("fixek: " + str(fix))
        elif "Sikeresen kicserélted" in line and "járművének olaját." in line:
            oilchange += 1
            print("oilchangek: " + str(oilchange))
        elif "Sikeresen felsegítetted a kiválasztott játékost" in line:
            asegit += 1
            print("asegitek: " + str(asegit))
        elif adminnick in line and "páncélját" in line:
            setarmor += 1
            print("setarmorok: " + str(setarmor))
        elif "[AdminDuty]: " + adminnick + " adminszolgálatba lépett." in line or "[Infobox]: " + adminnick + " adminszolgálatba lépett." in line:
            if "[" in line and "]" in line and "[Output]" in line:
                timeLine = line.split("[")
                timeLine = timeLine[1].split("]")
                timeLine = timeLine[0]
                print(timeLine)
                startTime = timeLine
                print("started duty: " + timeLine)
        elif "[AdminDuty]: " + adminnick + " kilépett az adminszolgálatból." in line:
            if "[" in line and "]" in line and "[Output]" in line:
                timeLine = line.split("[")
                timeLine = timeLine[1].split("]")
                timeLine = timeLine[0]
                print(timeLine)
            endTime = timeLine

            print("end: " + endTime)
            t1 = datetime.strptime(startTime, "%Y-%m-%d %H:%M:%S")
            t2 = datetime.strptime(endTime, "%Y-%m-%d %H:%M:%S")
            delta = t2 - t1
            dutyperc += delta.total_seconds()/60
            print("dutypercek: " + str(dutyperc))


        workbook = openpyxl.load_workbook("./admin.xlsx")
        sheet = workbook.active
        sheet["B4"] = dutyperc
        sheet["B6"] = asegit
        sheet["B8"] = setarmor

        sheet["E4"] = fix
        sheet["E6"] = oilchange

        workbook.save("./admin.xlsx")
